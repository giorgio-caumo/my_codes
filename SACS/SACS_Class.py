import sqlite3
import pandas as pd
import numpy as np
import math
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl import load_workbook
import os
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from pathlib import Path
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from fpdf import FPDF

from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4
from reportlab.platypus import Table as PDFTable
from reportlab.platypus import TableStyle
from reportlab.platypus import PageBreak, Paragraph, Table as PDFTable
from reportlab.lib import colors




class SACSReader:

    def __init__(self, index, database_file, selected_joints, selected_members, excel_file_path) -> None:
        """Create a SACS class reader

        Args:
            index (str): Unique name to distinguish if several SACS files have to be read
            database_file (Path): SACS database file
            selected_joints(list): selection of joints for filtering the data
            selected_members(list): selection of members for filtering the data
            excel_file_path(Path): file path where the excel has to be saved. Same location will be for the PDF

        Returns:
            Class: Class object
        """
        self.index = index
        self.database_file = database_file
        self.selected_joints = selected_joints
        self.selected_members = selected_members
        self.excel_file_path = excel_file_path

    def open_database(self, database_file: Path):
        """Open a connection to a SACS database

        Args:
            database_file (Path): Path to the databse

        Returns:
            Tuple: connection and cursor
        """
        conn = sqlite3.connect(database_file)
        cursor = conn.cursor()
        self.conn = conn
        self.cursor = cursor

        return conn, cursor

    def close_database(self):
        self.conn.close()

    def get_joint_name(self, cursor):
        """Obtain the Joint name as the user defined them. In the database the results are originally named with the joint ID
        which is not the user name convention

        Args:
            cursor

        Returns:
            df: Dataframe for mapping IDs
        """
        # Find the joint name ordered depending on the ID given by the db
        cursor.execute("SELECT ID, JointName FROM R_JOINT")
        # Fetch all rows from the query
        rows = cursor.fetchall()
        # Create a pandas DataFrame from the fetched data
        df = pd.DataFrame(rows, columns=['Joint ID', 'Joint Name'])
        # Make sure that all the elements are a string
        df['Joint Name'] = df['Joint Name'].astype(str)

        return df

    def get_member_name(cursor):
        ""Obtain the Member names 

        Args:
            cursor

        Returns:
            df: Dataframe for mapping Members
        """
        
        # Find the member name ordered depending on the ID given by the db
        cursor.execute("SELECT MemberName, ID FROM R_POSTMEMBERRESULTS")
        # Fetch all rows from the query
        rows = cursor.fetchall()
        # Create a pandas DataFrame from the fetched data
        df = pd.DataFrame(rows, columns=['Member Name', 'Member ID'])

        return df

    def get_fixed_joint_reactions(self, cursor, selected_joints=None) -> pd.DataFrame:
        """Get the fixed joint reaction

        Args:
            cursor (_type_): cursor to the SACS databse
            selected_joints (_type_, optional): Selection of Joints if the user wants specific ones

        Returns:
            None
        """
        # Get the selection if present
        selected_joints = self.selected_joints

        joint_names = self.get_joint_name(self.cursor)

        # Select all the results from fixed joint reaction
        cursor.execute(
            "SELECT LoadConditionName, JointID, ForceX, ForceY, ForceZ, MomentX,	MomentY, MomentZ FROM R_POSTJOINTREACTIONRESULTS")

        # Fetch all rows from the query
        rows = cursor.fetchall()

        # Create a pandas DataFrame from the fetched data
        df = pd.DataFrame(rows, columns=['Load Condition', 'Joint ID', 'Force X', 'Force Y', 'Force Z', 'Moment X',
                                         'Moment Y', 'Moment Z'])

        # Create a dictionary to map IDs to Names
        id_to_name = joint_names.set_index('Joint ID')['Joint Name'].to_dict()
        # Replace the ID column in df with the corresponding names
        df['Joint ID'] = df['Joint ID'].map(id_to_name)

        if selected_joints is not None:
            df_selected_joints = df[df['Joint ID'].isin(selected_joints)]
            self.fixed_joint_reactions = df_selected_joints
            
        else:
            self.fixed_joint_reactions = df
            

    def get_joint_spring_forces(self, cursor, selected_joints=None) -> pd.DataFrame:
        """Get the spring joint reactions

        Args:
            cursor (_type_): cursor to the SACS databse
            selected_joints (_type_, optional): Selection of Joints if the user wants specific ones

        Returns:
            None
        """
        # Get the selection if present
        selected_joints = self.selected_joints

        # Select the results from joint spring reaction
        joint_names = self.get_joint_name(self.cursor)
        cursor.execute(
            "SELECT LoadConditionName, JointID, GlobalFx, GlobalFy, GlobalFz, GlobalMx, GlobalMy, GlobalMz FROM R_POSTJOINTSPRINGREACTIONRESULTS")

        # Fetch all rows from the query
        rows = cursor.fetchall()

        # Create a pandas DataFrame from the fetched data
        df = pd.DataFrame(rows,
                          columns=['Load Condition', 'Joint ID', 'Force X', 'Force Y', 'Force Z', 'Moment X',
                                   'Moment Y', 'Moment Z'])

        # Create a dictionary to map IDs to Names
        id_to_name = joint_names.set_index('Joint ID')['Joint Name'].to_dict()

        # Replace the ID column in df with the corresponding names
        df['Joint ID'] = df['Joint ID'].map(id_to_name)

        if selected_joints is not None:
            df_selected_joints = df[df['Joint ID'].isin(selected_joints)]
            self.spring_forces = df_selected_joints
            # return df_selected_joints
        else:
            self.spring_forces = df
            # return df

    def get_member_end_forces(self, cursor,  selected_members=None) -> pd.DataFrame:
        """Get the fixed joint reaction

            Args:
                cursor (_type_): cursor to the SACS databse
                member_names (pd.Dataframe): _description_
                selected_members (_type_, optional): _description_. Defaults to None.

            Returns:
                _type_: _description_
        """
        # Get the selection if present
        selected_members = self.selected_members

        # Open the member end results file
        cursor.execute(
            "SELECT MemberName, MemberGroup, LoadConditionName, ForceXA, ForceYA, ForceZA, MomentXA, MomentYA, MomentZA FROM R_SOLVEMEMBERENDFORCESRESULTS")

        # Fetch all rows from the query
        rows = cursor.fetchall()

        # Create a pandas DataFrame from the fetched data
        df = pd.DataFrame(rows,
                          columns=['MemberName', 'MemberGroup', 'Load Condition', 'Force X', 'Force Y', 'Force Z',
                                    'Moment X', 'Moment Y', 'Moment Z'])
        # Extract 'xxxx' and 'yyyy' parts from Member Name, so that can be created a column with the corresponding end
        df[['xxxx', 'yyyy']] = df['MemberName'].str.split('-', expand=True)
        # Get the first Load Case name
        mask_LC = df.loc[0, 'Load Condition']
        counter = -1

        # Create 'Counter' and 'Member End' columns
        df['Counter'] = 0
        df['Member End'] = ''
        # Row by row, if the Load Condition is the same as the mask, increase the counter.
        # Based on the counter the name of the end is the initial part of the ending.
        for index, row in df.iterrows():
            load_case = row['Load Condition']

            # Check if the current Load Case matches the mask
            if load_case == mask_LC:
                counter += 1
                df.at[index, 'Counter'] = counter
                # Set 'Member End' based on whether the counter is even or odd
                df.at[index, 'Member End'] = df.at[index, 'xxxx'] if counter % 2 == 0 else df.at[index, 'yyyy']

            else:
                # If the Load Case doesn't match, save the current counter value
                df.at[index, 'Counter'] = counter
                # Set 'Member End' based on whether the counter is even or odd
                df.at[index, 'Member End'] = df.at[index, 'xxxx'] if counter % 2 == 0 else df.at[index, 'yyyy']

        # Drop intermediate columns if needed
        df = df.drop(['xxxx', 'yyyy', 'Counter'], axis=1)

        # Rearrange the columns
        df = df[
                ['MemberName', 'Member End', 'MemberGroup', 'Load Condition', 'Force X', 'Force Y', 'Force Z',
                 'Moment X',
                 'Moment Y', 'Moment Z']]

        if selected_members is not None:
            # Filter only the required members
            df_member_end = df[df['MemberName'].isin(selected_members)]
            self.member_end_forces = df_member_end
            # return df_member_end
        else:
            self.member_end_forces = df
            # return df

    # Function to find the maximum member UC
    def get_member_UC(self, cursor, selected_members=None)-> pd.DataFrame:
        """Get the fixed joint reaction

                        Args:
                            cursor (_type_): cursor to the SACS databse
                            joint_names (pd.Dataframe): _description_
                            selected_joints (_type_, optional): _description_. Defaults to None.

                        Returns:
                            _type_: _description_
                        """
        # Get the selection if present
        selected_members= self.selected_members # Execute a query to fetch columns A, B, and C from table2
        cursor.execute(
                "SELECT MemberName, MemberGroup, LoadConditionName, Distance, MemberLength, MaxUC FROM R_POSTMEMBERRESULTS")

        # Fetch all rows from the query
        rows = cursor.fetchall()

        # Create a pandas DataFrame from the fetched data
        df = pd.DataFrame(rows,
                        columns=['Member Name', 'Member Group', 'Load ConditionName', 'Distance', 'Member Length',
                                       'MaxUC'])

        if selected_members is not None:
            df_selected_members = df[df['Member Name'].isin(selected_members)]
            self.members_UC = df_selected_members
        else:
            self.members_UC = df

    def plot_forces(self, data, selection):
        print('Started Plotting')

        # Select the data
        if data == 'Fixed Joint Reactions':
            df = self.fixed_joint_reactions
        elif data == 'Spring Joint Forces':
            df = self.spring_forces

        # Select the axis
        if selection == 'X':
            axis = 'Force X'
        elif selection == 'Y':
            axis = 'Force Y'
        elif selection == 'Z':
            axis = 'Force Z'
        else:
            print("Error, use 'X', 'Y', 'Z'")

        # Step 1: Identify unique joint names
        joint_names = df['Joint ID'].unique()
        num_joints = len(joint_names)

        # Step 2: Determine optimal subplot grid
        rows = cols = math.ceil(math.sqrt(num_joints))

        # Step 3: Create subplots
        fig, axes = plt.subplots(rows, cols, figsize=(12, 12))
        axes = axes.flatten()

        for i, joint in enumerate(joint_names):
            joint_data = df[df['Joint ID'] == joint]
            bars = axes[i].bar(joint_data['Load Condition'], joint_data[axis])
            axes[i].set_title(f'Joint: {joint}')
            axes[i].set_xlabel('Load Case')
            axes[i].set_ylabel('Value')

            # Adding value labels on top of each bar
            for bar in bars:
                yval = bar.get_height()
                axes[i].text(bar.get_x() + bar.get_width() / 2, yval, round(yval, 2),
                        ha='center', va='bottom')

        # Hide any unused subplots if the grid is larger than needed
        for j in range(i + 1, rows * cols):
            fig.delaxes(axes[j])


        plt.tight_layout()
        plt.show()

    def get_data(self):
        conn, cursor = self.open_database(self.database_file)
        self.get_fixed_joint_reactions(cursor)
        self.get_joint_spring_forces(cursor)
        self.get_member_end_forces(cursor)
        self.get_member_UC(cursor)

        self.close_database()

    def save_results_to_excel(self, sheet_name, df):
        excel_file_path = self.excel_file_path

        if not os.path.isfile(excel_file_path):
            # If the file doesn't exist, save the changes to create a new Excel file
            wb = Workbook()
            # Remove the default "Sheet" since you don't want it
            default_sheet = wb["Sheet"]
            wb.remove(default_sheet)

            # Set the active sheet to the specified name
            ws = wb.create_sheet(sheet_name)
            wb.save(excel_file_path)
            print(f"Excel file '{excel_file_path}' created with '{sheet_name}' sheet.")

        # Load the existing Excel file
        wb = load_workbook(excel_file_path)

        # Check if the sheet exists
        if sheet_name in wb.sheetnames:
            # Get the existing sheet
            ws = wb[sheet_name]

            # Clear all existing rows in the sheet (including the header row)
            ws.delete_rows(1, ws.max_row)
            # for row in list(ws.iter_rows()):
            #     ws.delete_rows(idx=row[0].row, amount=1)

        else:
            # If the sheet doesn't exist, create a new one
            ws = wb.create_sheet(sheet_name)

        # Create title cell
        end_column = get_column_letter(len(df.columns))
        title_range = f'A1:{end_column}1'
        ws.merge_cells(title_range)
        # Set font to Times New Roman
        title_font = Font(name='Helvetica', bold=True)
        ws['A1'].font = title_font

        ws['A1'] = sheet_name
        ws['A1'].alignment = Alignment(horizontal='center')

        # ----------------------
        # Apply borders to all cells in the dataframe
        for row in ws.iter_rows(min_row=3, max_row=len(df) + 3, max_col=len(df.columns)):
            for cell in row:
                cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'),
                                         right=Side(style='thin'))
        # ----------------------
        # Write the dataframe to the worksheet, starting from cell A3
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False), start=3):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # ----------------------

        # Set column widths
        for i, column in enumerate(df, start=1):
            column_letter = get_column_letter(i)
            ws.column_dimensions[column_letter].width = len(str(column)) + 2
        # ----------------------

        # Save the changes to the existing Excel file
        wb.save(excel_file_path)
        print(f"Data updated in '{sheet_name}' sheet.")

    def export_to_excel(self, selection:str) -> None:
        excel_file_path = self.excel_file_path

        if selection == "Fixed Joint Reactions":
            df = self.fixed_joint_reactions
            sheet_name = f"{self.index} - FJR"
            self.save_results_to_excel(sheet_name= sheet_name, df=df)

        elif selection == 'Spring Joint Forces':
            df = self.spring_forces
            sheet_name = f"{self.index} - SJF"
            self.save_results_to_excel(sheet_name=sheet_name, df=df)

        elif selection == 'Member End Forces':
            df = self.member_end_forces
            sheet_name = f"{self.index} - MEF"
            self.save_results_to_excel(sheet_name=sheet_name, df=df)

        elif selection == 'Member UC':
            df = self.members_UC
            sheet_name = f"{self.index} - MUC"
            self.save_results_to_excel(sheet_name=sheet_name, df=df)

    def export_to_pdf(self):
        excel_file_path = self.excel_file_path
        pdf_file_path = os.path.splitext(excel_file_path)[0] + '.pdf'

        pdf = FPDF()
        # Add a page to the PDF
        pdf.add_page()
        # Set font for the PDF
        pdf.set_font("Arial", size=11)

        # Convert Excel file to PDF
        wb = load_workbook(excel_file_path)
        for sheet in wb.sheetnames:
            pdf.cell(200, 10, txt=sheet, ln=True)
            df = pd.read_excel(excel_file_path, sheet_name=sheet)
            for col in df.columns:
                pdf.cell(40, 10, str(col), ln=False)
            pdf.ln()
            for index, row in df.iterrows():
                for col in df.columns:
                    pdf.cell(40, 10, str(row[col]), ln=False)
                pdf.ln()

        # Save the PDF
        pdf.output(pdf_file_path)
        print(f"PDF file '{pdf_file_path}' created.")

    def export_to_pdf_styled(self):
        excel_file_path = self.excel_file_path
        pdf_file_path = os.path.splitext(excel_file_path)[0] + '.pdf'

        # Create a PDF document
        doc = SimpleDocTemplate(pdf_file_path, pagesize=letter,
                                # leftMargin=36, rightMargin=36,
                                topMargin=30, bottomMargin=30)
        elements = []
        # Get the styles
        styles = getSampleStyleSheet()

        # Read the data from the Excel file
        wb = load_workbook(excel_file_path, read_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            data = []

            # Read the data from the worksheet
            for row in ws.iter_rows(values_only=True):
                formatted_row = [round(cell, 2) if isinstance(cell, float) else cell for cell in row]
                data.append(formatted_row)

            # Calculate column widths based on the maximum length of data in each column
            if len(data[0]) <10:
                col_widths = [max(len(str(cell)) for cell in col) *8 for col in zip(*data)]
            else:
                col_widths = [max(len(str(cell)) for cell in col) * 6 for col in zip(*data)]

            # Create a PDF table from the data
            pdf_table_data = [list(row) for row in data]
            pdf_table = PDFTable(pdf_table_data)

            # Apply the styles
            pdf_table.setStyle([
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('BACKGROUND', (0, 2), (-1, 2), colors.gray),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 3), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ])

            # Set column widths
            pdf_table._argW = col_widths

            # Add the PDF table to the elements list
            elements.append(pdf_table)

            # Add a page break after each sheet except the last one
            # if sheet != wb.sheetnames[-1]:
            #     elements.append(PageBreak())

            # # Add a page break after each sheet
            elements.append(PageBreak())

        # Build the PDF document
        doc.build(elements)

        # Remove the last page break
        # elements.pop()

        print(f"PDF file '{pdf_file_path}' created.")

    def get_max(self, column_name, columns_to_keep, title, df) -> pd.DataFrame:
        # Find the maximum value in a specific column (e.g., 'Force X')
        max_value = df[column_name].max()
        # Get the entire row where the maximum value is located
        rows_with_max_value = df[df[column_name] == max_value]

        # Filter the DataFrame to keep only the necessary columns
        row_max = rows_with_max_value.filter(items=columns_to_keep)

        # Concatenate the title row with the row_max DataFrame
        result_df = pd.concat([title, row_max], ignore_index=True)

        return result_df

    def get_min(self, column_name, columns_to_keep, title, df) -> pd.DataFrame:
        # Find the maximum value in a specific column (e.g., 'Force X')
        min_value = df[column_name].min()
        # Get the entire row where the maximum value is located
        rows_with_min_value = df[df[column_name] == min_value]

        # Filter the DataFrame to keep only the necessary columns
        row_min = rows_with_min_value.filter(items=columns_to_keep)

        # Concatenate the title row with the row_max DataFrame
        result_df = pd.concat([title, row_min], ignore_index=True)

        return result_df

    def post_process_fixed_reactions(self, selection = str, direction = str, selected_joints = None)-> pd.DataFrame:
        print(f"Looking for {selection} in {direction}...")
        df = self.fixed_joint_reactions

        #Pick the right direction
        if direction == 'X':
            column_name = 'Force X'
        elif direction == 'Y':
            column_name = 'Force Y'
        elif direction == 'Z':
            column_name = 'Force Z'
        elif direction == 'Mx':
            column_name = 'Moment X'
        elif direction == 'My':
            column_name = 'Moment Y'
        elif direction == 'Mz':
            column_name = 'Moment Z'
        else:
            raise ValueError("Invalid direction. Use 'X', 'Y', or 'Z' for forces, 'Mx', 'My', 'Mz' for moments.")

        #Select only the necessary joints
        if selected_joints is not None:
            filtered_df = df[df['Joint ID'].isin(selected_joints)]
        else:
            filtered_df = df

        # Create a DataFrame for the title row
        title_row = pd.DataFrame([[f"{selection} - {direction}"]], columns=['Title'])

        # List of columns to keep
        columns_to_keep = ['Load Condition', 'Joint ID', column_name]

        if selection == 'Max':

            result_df = self.get_max(column_name = column_name, columns_to_keep=columns_to_keep, title=title_row, df=filtered_df)
            # # Find the maximum value in a specific column (e.g., 'Force X')
            # max_value = filtered_df[column_name].max()
            # # Get the entire row where the maximum value is located
            # rows_with_max_value = filtered_df[filtered_df[column_name] == max_value]
            #
            # # Filter the DataFrame to keep only the necessary columns
            # row_max = rows_with_max_value.filter(items=columns_to_keep)
            #
            # # Concatenate the title row with the row_max DataFrame
            # result_df = pd.concat([title_row, row_max], ignore_index=True)

        elif selection == 'Min':

            result_df = self.get_min(column_name=column_name, columns_to_keep=columns_to_keep, title=title_row,
                                     df=filtered_df)
            # # Find the maximum value in a specific column (e.g., 'Force X')
            # min_value = filtered_df[column_name].min()
            # # Get the entire row where the maximum value is located
            # rows_with_min_value = filtered_df[filtered_df[column_name] == min_value]
            #
            # # Filter the DataFrame to keep only the necessary columns
            # row_min = rows_with_min_value.filter(items=columns_to_keep)
            #
            # # Concatenate the title row with the row_max DataFrame
            # result_df = pd.concat([title_row, row_min], ignore_index=True)

        print(f"{selection} found")
        return result_df

    def post_process_spring_forces(self, selection = str, direction = str, selected_joints = None)-> pd.DataFrame:
        print(f"Looking for {selection} in {direction}...")
        df = self.spring_forces

        #Pick the right direction
        if direction == 'X':
            column_name = 'Force X'
        elif direction == 'Y':
            column_name = 'Force Y'
        elif direction == 'Z':
            column_name = 'Force Z'
        elif direction == 'Mx':
            column_name = 'Moment X'
        elif direction == 'My':
            column_name = 'Moment Y'
        elif direction == 'Mz':
            column_name = 'Moment Z'
        else:
            raise ValueError("Invalid direction. Use 'X', 'Y', or 'Z' for forces, 'Mx', 'My', 'Mz' for moments.")

        #Select only the necessary joints
        if selected_joints is not None:
            filtered_df = df[df['Joint ID'].isin(selected_joints)]
        else:
            filtered_df = df

        # Create a DataFrame for the title row
        title_row = pd.DataFrame([[f"{selection} - {direction}"]], columns=['Title'])

        # List of columns to keep
        columns_to_keep = ['Load Condition', 'Joint ID', column_name]

        if selection == 'Max':
            result_df = self.get_max(column_name=column_name, columns_to_keep=columns_to_keep, title=title_row,
                                     df=filtered_df)

        elif selection == 'Min':
            result_df = self.get_min(column_name=column_name, columns_to_keep=columns_to_keep, title=title_row,
                                     df=filtered_df)

        print(f"{selection} found")
        return result_df
    def post_process_member_forces(self, selection=str, direction=str, selected_members=None)-> pd.DataFrame:
        print(f"Looking for {selection} in {direction}...")
        df = self.member_end_forces

        # Pick the right direction
        if direction == 'X':
            column_name = 'Force X'
        elif direction == 'Y':
            column_name = 'Force Y'
        elif direction == 'Z':
            column_name = 'Force Z'
        elif direction == 'Mx':
            column_name = 'Moment X'
        elif direction == 'My':
            column_name = 'Moment Y'
        elif direction == 'Mz':
            column_name = 'Moment Z'
        else:
            raise ValueError("Invalid direction. Use 'X', 'Y', or 'Z' for forces, 'Mx', 'My', 'Mz' for moments.")

        # Select only the necessary joints
        if selected_members is not None:
            filtered_df = df[df['MemberName'].isin(selected_members)]
        else:
            filtered_df = df

        # Create a DataFrame for the title row
        title_row = pd.DataFrame([[f"{selection} - {direction}"]], columns=['Title'])

        # List of columns to keep
        columns_to_keep = ['MemberName', 'Load Condition', column_name]

        if selection == 'Max':
            result_df = self.get_max(column_name=column_name, columns_to_keep=columns_to_keep, title=title_row,
                                     df=filtered_df)

        elif selection == 'Min':
            result_df = self.get_min(column_name=column_name, columns_to_keep=columns_to_keep, title=title_row,
                                     df=filtered_df)

        print(f"{selection} found")
        return result_df


if __name__ == "__main__":

    # THISDIR = Path(__file__).parent
    excel_file = os.path.join(THISDIR, 'SACS output_TEST.xlsx')

    database_file = [r'Operational\sacsdb.post',
                 r'Spare\sacsdb.post']

    selected_joints = None
    selected_members = ['SD01-TOG', 'SD02-TOG', 'SD03-TOG', 'SD04-TOG', 'SD05-TOG', 'SD06-TOG', 'SD07-TOG', 'SD08-TOG',
                            # For the moment Detail 7
                    'SD01-D01', 'SD02-D02', 'SD03-D03', 'SD04-D04', 'SD05-D05', 'SD06-D06', 'SD07-D07', 'SD08-D08'
                            ]
    index = ["Operational", "Spare"]


